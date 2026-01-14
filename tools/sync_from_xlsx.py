"""
Sync site data/pages from SheTech_Career_Map.xlsx.

- Reads the Excel sheet into a list of careers
- Writes careers-data.js (used by landing page)
- Generates/updates careers/*.html for every career in the sheet
- Removes careers/*.html that are no longer in the sheet

Run:
  python tools/sync_from_xlsx.py
"""

from __future__ import annotations

from pathlib import Path
import json
import re
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "SheTech_Career_Map.xlsx"
CAREERS_DATA_JS = ROOT / "careers-data.js"
CAREERS_DIR = ROOT / "careers"


def slugify(title: str) -> str:
    s = title.strip().lower()
    s = s.replace("&", " and ")
    s = s.replace("/", " ")
    s = re.sub(r"[^a-z0-9\s-]", "", s)
    s = re.sub(r"\s+", "-", s).strip("-")
    s = re.sub(r"-{2,}", "-", s)
    return s


def infer_category(title: str) -> str:
    t = title.lower()
    if any(k in t for k in ["data scientist", "ai researcher", "ai", "ml", "prompt", "genomic"]):
        return "data"
    if "engineer" in t or "uav" in t or "drone" in t:
        return "engineering"
    if any(
        k in t
        for k in [
            "designer",
            "animator",
            "architect",
            "choreographer",
            "marketer",
            "producer",
            "production",
            "graphics",
            "entrepreneur",
        ]
    ):
        return "design"
    if any(k in t for k in ["scientist", "biologist", "chemist", "astronom", "climate", "forensic", "genetic", "sustainability", "astrophys"]):
        return "science"
    return "technology"


def image_query(title: str, img_desc: str) -> str:
    base = "woman"
    text = (img_desc or "").strip() or title
    text = re.sub(r"\b(woman|female|young)\b", "", text, flags=re.I)
    text = re.sub(r"[^a-zA-Z0-9\s]", " ", text)
    words = [w.strip().lower() for w in text.split() if w.strip()]
    keywords: list[str] = []
    for w in words:
        if w in {"in", "a", "the", "with", "and", "or", "of", "to", "on", "front", "room", "near"}:
            continue
        if w not in keywords:
            keywords.append(w)
        if len(keywords) >= 5:
            break
    if not keywords:
        keywords = [w for w in re.sub(r"[^a-zA-Z0-9\s]", " ", title).lower().split()][:3]
    return ",".join([base] + keywords)


def choose_article(title: str) -> str:
    t = (title or "").strip()
    if not t:
        return "a"
    first = (re.match(r"^[A-Za-z0-9]+", t) or [""])[0]
    upper = first.upper()
    if upper in {"UX", "UI", "EU"}:
        return "a"
    vowel_sound_letters = set("AEFHILMNORSX")
    is_acronym = bool(re.match(r"^[A-Z0-9]{2,4}$", first))
    if is_acronym:
        return "an" if upper[:1] in vowel_sound_letters else "a"
    return "an" if re.match(r"^[AEIOU]", first, re.I) else "a"


def build_pathway_table_lists(items: list[str]) -> str:
    if not items:
        return "<li class=\"muted\">(Add items)</li>"
    return "".join(f"<li>{i}</li>" for i in items)


PAGE_TMPL = """<!doctype html>
<html lang="en">
  <head>
    <meta charset="utf-8" />
    <meta name="viewport" content="width=device-width, initial-scale=1" />
    <title>Launch Your Future as {article} {title} — SheTech Pathways</title>
    <meta name="description" content="{meta_desc}" />
    <link rel="stylesheet" href="../styles.css" />
  </head>
  <body>
    <header class="site-header">
      <div class="container header-inner">
        <a class="brand" href="../index.html" aria-label="Back to landing page">
          <img class="brand-mark" src="../assets/shetech_logo_white.png" alt="SheTech logo" />
          <span class="brand-text">Pathways</span>
        </a>
        <nav class="nav" aria-label="Primary">
          <a href="../index.html#careers">Careers</a>
          <a href="../qr-sheet.html">QR sheet</a>
        </nav>
      </div>
    </header>

    <main class="page">
      <div class="container">
        <header class="page-header">
          <p class="breadcrumbs">
            <a href="../index.html">Home</a> <span aria-hidden="true">/</span>
            <a href="../index.html#careers">Careers</a> <span aria-hidden="true">/</span>
            {title}
          </p>
          <h1 class="page-title">Launch Your Future as {article} {title}</h1>
          <p class="page-subtitle">{desc}</p>
        </header>

        <div class="content-grid">
          <div>
            <div class="career-hero" data-career-hero data-slug="{slug}" hidden>
              <img class="career-hero-img" alt="{title} hero image" loading="lazy" decoding="async" />
            </div>

            <!-- Google Doc-driven panels injected here -->
            <div id="docSections"></div>
          </div>

          <aside class="side-stack">
            <section class="panel">
              <h2>Share this pathway</h2>
              <div class="qr-box">
                <img alt="QR code for {title} pathway" width="200" height="200" data-qr data-path="./{slug}.html" />
                <p class="muted small">Host this site (so QR codes open the correct pages when scanned).</p>
              </div>
              <div class="side-actions">
                <a class="btn btn-primary" href="../index.html#careers">All careers</a>
                <button class="btn btn-ghost" type="button" data-copy-link data-path="./{slug}.html">Copy link</button>
              </div>
            </section>
          </aside>
        </div>
      </div>
    </main>

    <footer class="site-footer">
      <div class="container footer-inner">
        <div class="footer-left">
          <img class="footer-mark" src="../assets/shetech_logo_teal.png" alt="SheTech logo" loading="lazy" />
          <div>
            <p class="footer-title">SheTech Pathways</p>
            <p class="muted small">Explore more careers on the landing page.</p>
          </div>
        </div>
        <div class="footer-right">
          <a href="../index.html#careers">Careers</a>
        </div>
      </div>
    </footer>

    <script src="../careers-data.js"></script>
    <script src="../doc-media.js"></script>
    <script src="../script.js"></script>
  </body>
</html>
"""


def main() -> int:
    if not XLSX.exists():
        print(f"Missing {XLSX}")
        return 2

    CAREERS_DIR.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(XLSX, data_only=True)
    ws = wb[wb.sheetnames[0]]

    header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    header = [str(x).strip() if x is not None else "" for x in header]
    col = {name: i + 1 for i, name in enumerate(header) if name}

    required = ["Poster Title", "Description"]
    missing = [c for c in required if c not in col]
    if missing:
        print(f"Missing expected columns: {missing}")
        return 2

    items = []
    for r in range(2, ws.max_row + 1):
        title = ws.cell(r, col["Poster Title"]).value
        if not title or not str(title).strip():
            continue
        title = str(title).strip()
        desc = str(ws.cell(r, col["Description"]).value or "").strip()

        slug = slugify(title)
        items.append(
            {
                "title": title,
                "slug": slug,
                "category": infer_category(title),
                "description": desc,
                "highSchool": [],
                "college": [],
                "career": [],
                "imageDescription": "",
                "imageQuery": image_query(title, ""),
                "resources": [],
            }
        )

    CAREERS_DATA_JS.write_text("window.SHT_CAREERS = " + json.dumps(items, indent=2) + ";\n", encoding="utf-8")

    # Generate pages
    required_slugs = set()
    for c in items:
        title = c["title"]
        slug = c["slug"]
        required_slugs.add(slug)
        article = choose_article(title)
        meta_desc = c["description"] or f"Launch your future as {article} {title}—explore high school courses, college majors, and career roles."
        html = PAGE_TMPL.format(
            title=title,
            slug=slug,
            article=article,
            desc=c["description"],
            meta_desc=meta_desc.replace('"', "&quot;"),
            img_desc=c["imageDescription"] or title,
        )
        (CAREERS_DIR / f"{slug}.html").write_text(html, encoding="utf-8")

    # Remove pages no longer present
    for f in CAREERS_DIR.glob("*.html"):
        if f.stem not in required_slugs:
            f.unlink()

    print(f"Synced {len(items)} careers, wrote careers-data.js, and regenerated career pages.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())


